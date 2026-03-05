"""Generate the 7-sheet Excel Output Pack from a run_analysis() result dict."""
from __future__ import annotations

from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Sheet definitions: (sheet_key, sheet_title, columns)
# columns = list of (header_label, data_key_or_callable)
# ---------------------------------------------------------------------------

def _join_list(val: Any) -> str:
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val) if val is not None else ""


_SHEETS: List[tuple] = [
    (
        "executive_scorecard",
        "Executive Scorecard",
        [
            ("Run ID",              lambda r: r.get("run_id", "")),
            ("Generated At",        lambda r: r.get("generated_at", "")),
            ("Reports Scanned",     lambda r: r.get("reports_scanned", 0)),
            ("Datasets Extracted",  lambda r: r.get("datasets_extracted", 0)),
            ("KPIs Inferred",       lambda r: r.get("kpis_inferred", 0)),
            ("Clusters Formed",     lambda r: r.get("clusters_formed", 0)),
            ("Drift Findings",      lambda r: r.get("drift_findings_count", 0)),
            ("P1 Recommendations",  lambda r: r.get("p1_recommendations", 0)),
            ("Highest Severity",    lambda r: r.get("highest_severity", 0.0)),
        ],
    ),
    (
        "report_inventory",
        "Report Inventory",
        [
            ("Report ID",       lambda r: r.get("report_id", "")),
            ("Name",            lambda r: r.get("name", "")),
            ("Folder",          lambda r: r.get("folder", "")),
            ("Owner",           lambda r: r.get("owner") or ""),
            ("Last Modified",   lambda r: r.get("last_modified") or ""),
            ("Datasets",        lambda r: r.get("dataset_count", 0)),
            ("Primary Domain",  lambda r: r.get("primary_domain") or ""),
        ],
    ),
    (
        "kpi_dictionary",
        "KPI Dictionary",
        [
            ("KPI Name",            lambda r: r.get("kpi_name", "")),
            ("Inferred Definition", lambda r: r.get("inferred_definition", "")),
            ("Grain",               lambda r: r.get("grain", "")),
            ("Primary Source",      lambda r: r.get("primary_source", "")),
            ("Computed In Layer",   lambda r: r.get("computed_in_layer", "")),
            ("Found In Reports",    lambda r: _join_list(r.get("found_in_reports", []))),
            ("Risk Notes",          lambda r: r.get("risk_notes", "")),
        ],
    ),
    (
        "clusters",
        "Metric Clusters",
        [
            ("Cluster ID",      lambda r: r.get("cluster_id", "")),
            ("Metric Intent",   lambda r: r.get("metric_intent") or ""),
            ("Members",         lambda r: _join_list(r.get("members", []))),
            ("Confidence",      lambda r: r.get("confidence_score", 0.0)),
            ("Duplicate Count", lambda r: r.get("duplicate_count", 0)),
            ("Reports",         lambda r: _join_list(r.get("reports", []))),
        ],
    ),
    (
        "drift_findings",
        "Drift Findings",
        [
            ("Finding ID",          lambda r: r.get("finding_id", "")),
            ("Cluster ID",          lambda r: r.get("cluster_id", "")),
            ("KPI Name",            lambda r: r.get("kpi_name") or ""),
            ("Drift Type",          lambda r: r.get("drift_type", "")),
            ("Severity",            lambda r: r.get("severity_score", 0.0)),
            ("Explanation",         lambda r: r.get("explanation", "")),
            ("Impact",              lambda r: r.get("impact") or ""),
            ("Recommendation",      lambda r: r.get("recommendation_text") or ""),
            ("Impacted Reports",    lambda r: _join_list(r.get("impacted_reports", []))),
        ],
    ),
    (
        "recommendations",
        "Recommendations Backlog",
        [
            ("Recommendation ID",   lambda r: r.get("recommendation_id", "")),
            ("Cluster ID",          lambda r: r.get("cluster_id", "")),
            ("Area",                lambda r: r.get("area") or ""),
            ("Action Type",         lambda r: r.get("action_type", "")),
            ("Priority Label",      lambda r: r.get("priority_label") or ""),
            ("Priority Score",      lambda r: r.get("priority_score", 0.0)),
            ("Rationale",           lambda r: r.get("rationale", "")),
            ("Target Layer",        lambda r: r.get("target_layer", "")),
            ("Owner",               lambda r: r.get("owner") or ""),
            ("Impacted Reports",    lambda r: _join_list(r.get("impacted_reports", []))),
        ],
    ),
    (
        "how_to_read",
        "How to Read",
        [
            ("Section",  lambda r: r.get("section", "")),
            ("Content",  lambda r: r.get("content", "")),
        ],
    ),
]


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0284C7")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_title_style(cell) -> None:
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0F172A")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_scorecard_sheet(ws, data: dict, run_id: str, generated_at: str) -> None:
    """Executive Scorecard is a single record rendered as key-value pairs."""
    _write_title(ws, "Executive Scorecard", run_id, generated_at, col_count=2)
    ws.append(["Field", "Value"])
    for cell in ws[ws.max_row]:
        _apply_header_style(cell)
    for header, accessor in _SHEETS[0][2]:
        ws.append([header, accessor(data)])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def _write_title(ws, title: str, run_id: str, generated_at: str, col_count: int) -> None:
    last_col = get_column_letter(col_count)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    _apply_title_style(ws["A1"])
    ws.row_dimensions[1].height = 28

    meta_str = f"Run: {run_id}  |  Generated: {generated_at}"
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = meta_str
    ws["A2"].font = Font(italic=True, color="64748B")
    ws.row_dimensions[2].height = 16

    # Blank separator
    ws.append([])


def _write_data_sheet(ws, title: str, rows: list, columns: list, run_id: str, generated_at: str) -> None:
    col_count = max(len(columns), 1)
    _write_title(ws, title, run_id, generated_at, col_count)

    headers = [col[0] for col in columns]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        _apply_header_style(cell)

    for row in rows:
        ws.append([accessor(row) for _, accessor in columns])

    # Auto-width (capped at 60)
    for i, (header, _) in enumerate(columns, start=1):
        col_letter = get_column_letter(i)
        max_len = max(
            len(header),
            *(len(str(accessor(row))) for row in rows for _, accessor in [columns[i - 1]])
            if rows else [0],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def generate_excel_pack(analysis_result: dict, output_path: str) -> None:
    """Write a 7-sheet Excel Output Pack to *output_path*.

    Requires openpyxl to be installed.  Raises RuntimeError if unavailable.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    scorecard = analysis_result.get("executive_scorecard", {})
    run_id = str(scorecard.get("run_id", "unknown"))
    generated_at = str(scorecard.get("generated_at", ""))

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    sheet_name_map = {
        "executive_scorecard": "Executive_Scorecard",
        "report_inventory":    "Report_Inventory",
        "kpi_dictionary":      "KPI_Dictionary",
        "clusters":            "Metric_Clusters",
        "drift_findings":      "Drift_Findings",
        "recommendations":     "Recommendations_Backlog",
        "how_to_read":         "How_to_Read",
    }

    for sheet_key, sheet_title, columns in _SHEETS:
        ws = wb.create_sheet(title=sheet_name_map[sheet_key])
        data = analysis_result.get(sheet_key, {} if sheet_key == "executive_scorecard" else [])

        if sheet_key == "executive_scorecard":
            _write_scorecard_sheet(ws, data, run_id, generated_at)
        else:
            rows = data if isinstance(data, list) else []
            _write_data_sheet(ws, sheet_title, rows, columns, run_id, generated_at)

    wb.save(output_path)
